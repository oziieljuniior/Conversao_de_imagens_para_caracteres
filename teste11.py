from PIL import Image
from pytesseract import pytesseract
import cv2 as cv
#import numpy as np 
from matplotlib import pyplot as plt
from openpyxl import Workbook

#arquivo_excel = Workbook()
#planilha1 = arquivo_excel.active
#planilha1.title = "Amostra6"
#print(arquivo_excel.sheetnames)
#planilha1['A1'] = "Odds"
#planilha1['B1'] = "Total de apostadores"
#planilha1['C1'] = "Online"

#C:\Users\ozcon\Documents\tesseract-ocr-w64-setup-v5.1.0.20220510
#C:\Program Files\Tesseract-OCR\tesseract.exe
#pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
wb = Workbook()
ws = wb.active
wb.title = "Amostra"
i = 0
r = 1
basewidth = 450
while i <= 301:
    i = str(i)
    r = str(r)
    aa = i  + "img.png"
    #bb = i + "T_Apostadores.png"
    #cc = i + "T_Online.png"
    
    a = Image.open(aa)
    #a = ImageEnhance.Contrast(a)
    #b = Image.open(bb)
    #c = Image.open(cc)
        
    
    #w0_percent = (basewidth/float(b.size[0]))
    #h0size = int((float(b.size[1])*float(w0_percent)))
    #b = b.resize((basewidth,h0size),Image.Resampling.LANCZOS)
    #b.save("e.png")
    #e = Image.open("e.png")
    #e = ImageEnhance.Color(e)
    
    
    #w1_percent = (basewidth/float(c.size[0]))
    #h1size = int(float(c.size[1])*float(w1_percent))
    #c = c.resize((basewidth,h1size),Image.Resampling.LANCZOS)
    #c.save("f.png")
    #f = Image.open("f.png")
    #f = ImageEnhance.Color(f) 
    
    image1 = cv.imread(aa)
    gray = cv.cvtColor(image1, cv.COLOR_BGR2GRAY)
    blur = cv.GaussianBlur(gray, (3,3), 0)
    thresh = cv.threshold(blur,0,255,cv.THRESH_BINARY_INV + cv.THRESH_OTSU)[1]
    #thresh1 = cv.threshold(thresh, 44,54,cv.THRESH_TRUNC)
    
    kernel = cv.getStructuringElement(cv.MORPH_RECT,(3,3))
    opening = cv.morphologyEx(thresh, cv.MORPH_OPEN, kernel, iterations = 5)
    invert = 255 - opening
    
    #image2 = cv.imread(bb)
    #gray0 = cv.cvtColor(image2, cv.COLOR_BGR2GRAY)
    #blur0 = cv.GaussianBlur(gray0, (3,3), 0)
    #thresh0 = cv.threshold(blur0,0,255,cv.THRESH_BINARY_INV + cv.THRESH_OTSU)[1]
    
    #kernel0 = cv.getStructuringElement(cv.MORPH_RECT,(3,3))
    #opening0 = cv.morphologyEx(thresh0, cv.MORPH_OPEN, kernel0, iterations = 1)
    #invert0 = 255 - opening0

    plt.imshow(invert),plt.colorbar(),plt.show()   

    g = pytesseract.image_to_string(invert, config = "--psm 6")
    #h = pytesseract.image_to_string(invert0, config = "--psm 6")
    #k = pytesseract.image_to_string(cv.cvtColor(np.array(f), cv.COLOR_BGR2GRAY), config = "--psm 6")
    
    print(type(g))
    #print(type(h))
    #print(type(k))
    
    o = 'A' + r
    #p = 'B' + r
    #q = 'C' + r
    
    bad = ["\n\x0c","x","\n","X","/", ")\ ", "Oe","%","Molo)","A","y",")", "We","@",""]
    
    for s in bad:
        g = g.replace(s,'')
     #   h = h.replace(s,'')
    #    k = k.replace(s,'')
    bad0 = [" "]
    for ss in bad0:
        g = g.replace(ss,'.')
     #   k = str(k)
  #  C1 = "A"+k
   # print(C1)
    #arquivo_excel = Workbook()
    #planilha1 = arquivo_excel.active
    #planilha1["C1"] = "a0"
    #arquivo_excel.save("Amostra6")
    #k = int(k)
    #k = k + 1
   
    print(o)
    ws[o] = g
    #ws[p] = h
    #ws[q] = k
    
    
    n = (i,g)
    l = list(n)
    print(l)
    array = []
    array.append(n)
    print(array)
    i = int(i)
    i = i + 1
    wb.save("Amostra.xlsx")
    print("Converteu")
    r = int(r)
    r = r + 1